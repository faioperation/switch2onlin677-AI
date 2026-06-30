"""
utils/excel_styles.py
=====================
Shared Excel styling constants and column definitions for the Dhifaf Baghdad
product export system.

All style objects are module-level singletons — created once on import and
reused across every row written. This matters for write_only workbooks where
hundreds of thousands of cells might reference the same style.
"""
from __future__ import annotations

from typing import Callable, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Color palette (hex strings, no leading #) ─────────────────────────────────
C_HEADER_BG   = "1E3A5F"   # Dark navy          — column header row
C_HEADER_FG   = "FFFFFF"   # White              — header text
C_TITLE_BG    = "0D2137"   # Near-black navy    — report title
C_TITLE_FG    = "F5C518"   # Gold               — title accent text
C_META_BG     = "EBF0F7"   # Light blue-grey    — metadata rows 2-4
C_SEP_BG      = "CDD8E8"   # Steel blue         — separator row 5
C_ROW_ODD     = "F4F7FB"   # Very light stripe  — odd data rows
C_ROW_EVEN    = "FFFFFF"   # White              — even data rows
C_IN_STOCK    = "E8F5E9"   # Soft green         — qty > 5
C_LOW_STOCK   = "FFF8E1"   # Amber              — qty 1-5 (low stock warning)
C_OUT_STOCK   = "FFEBEE"   # Soft red           — qty 0
C_BORDER      = "BCC8D8"   # Steel blue-grey    — cell borders

# ── Border sides ──────────────────────────────────────────────────────────────
_THIN   = Side(border_style="thin",   color=C_BORDER)
_MEDIUM = Side(border_style="medium", color="8FA8C8")

BORDER_CELL   = Border(left=_THIN,   right=_THIN,   top=_THIN,   bottom=_THIN)
BORDER_HEADER = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)
BORDER_NONE   = Border()

# ── Alignments ────────────────────────────────────────────────────────────────
ALIGN_TITLE   = Alignment(horizontal="left",   vertical="center", wrap_text=False, indent=1)
ALIGN_META    = Alignment(horizontal="left",   vertical="center", wrap_text=False, indent=1)
ALIGN_HEADER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ALIGN_RIGHT   = Alignment(horizontal="right",  vertical="center", wrap_text=False)

# ── Fills ─────────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

FILL_TITLE     = _fill(C_TITLE_BG)
FILL_HEADER    = _fill(C_HEADER_BG)
FILL_META      = _fill(C_META_BG)
FILL_SEP       = _fill(C_SEP_BG)
FILL_ROW_ODD   = _fill(C_ROW_ODD)
FILL_ROW_EVEN  = _fill(C_ROW_EVEN)
FILL_IN_STOCK  = _fill(C_IN_STOCK)
FILL_LOW_STOCK = _fill(C_LOW_STOCK)
FILL_OUT_STOCK = _fill(C_OUT_STOCK)

# ── Fonts ─────────────────────────────────────────────────────────────────────
FONT_TITLE    = Font(name="Calibri", bold=True,  size=14, color=C_TITLE_FG)
FONT_META_KEY = Font(name="Calibri", bold=True,  size=10, color="2D3748")
FONT_META_VAL = Font(name="Calibri", bold=False, size=10, color="4A5568")
FONT_HEADER   = Font(name="Calibri", bold=True,  size=10, color=C_HEADER_FG)
FONT_DATA     = Font(name="Calibri", bold=False, size=10, color="1A202C")

# ── Number formats ────────────────────────────────────────────────────────────
FMT_TEXT    = "@"
FMT_INT     = "#,##0"
FMT_USD     = '"$"#,##0.00'
FMT_IQD     = '#,##0" IQD"'
FMT_DECIMAL = "0.000000"
FMT_PCT     = "0.00"

# ── Column definitions ────────────────────────────────────────────────────────
# Tuple layout: (header_label, key_or_callable, col_width, number_format)
#
# key_or_callable:
#   "__row_num__"  → sequential 1-based export row number
#   str            → direct key lookup in the serialized row dict
#   callable       → called with the row dict; its return value is the cell value
#
# col_width is in Excel "character units" (approximately 8px each).
#
EXPORT_COLUMNS: list[tuple[str, str | Callable, int, Optional[str]]] = [
    # ── Identity ─────────────────────────────────────────────────────────────
    ("#",                  "__row_num__",                                      5,  FMT_INT),
    ("Barcode",            "barcode",                                          18, FMT_TEXT),
    ("Item Code",          "item_code",                                        15, FMT_TEXT),
    ("SAP Product ID",     "sap_product_id",                                   18, FMT_TEXT),
    ("Product Name",       "item_name",                                        36, FMT_TEXT),
    # ── Relations ────────────────────────────────────────────────────────────
    ("Brand",              lambda r: (r.get("brand")       or {}).get("name"), 18, FMT_TEXT),
    ("Category",           lambda r: (r.get("category")    or {}).get("name"), 18, FMT_TEXT),
    ("Subcategory",        lambda r: (r.get("subcategory") or {}).get("name"), 18, FMT_TEXT),
    # ── Display ──────────────────────────────────────────────────────────────
    ("Description",        "description",                                      40, FMT_TEXT),
    ("Image URL",          "image_url",                                        30, FMT_TEXT),
    # ── AI / Search ──────────────────────────────────────────────────────────
    ("Skin Type",          "skin_type",                                        14, FMT_TEXT),
    ("Concerns",           lambda r: ", ".join(r.get("concerns") or []),       22, FMT_TEXT),
    ("Tags",               lambda r: ", ".join(r.get("tags")     or []),       20, FMT_TEXT),
    # ── Pricing ──────────────────────────────────────────────────────────────
    ("Price (USD)",        "price",                                            14, FMT_USD),
    ("Price (IQD)",        "price_iqd",                                        16, FMT_IQD),
    ("Available Qty",      "available_qty",                                    14, FMT_INT),
    ("Stock Status",       "stock_status",                                     14, FMT_TEXT),
    # ── Classification ───────────────────────────────────────────────────────
    ("Price Tier",         "price_tier",                                       12, FMT_TEXT),
    ("Brand Family",       "brand_family",                                     18, FMT_TEXT),
    ("Product Status",     "product_status",                                   14, FMT_TEXT),
    # ── Recommendation flags ─────────────────────────────────────────────────
    ("Best Selling",       lambda r: "Yes" if r.get("is_best_selling")    else "No", 13, FMT_TEXT),
    ("New Arrival",        lambda r: "Yes" if r.get("is_new_arrival")     else "No", 13, FMT_TEXT),
    ("Recommended",        lambda r: "Yes" if r.get("is_recommended")     else "No", 13, FMT_TEXT),
    ("COD Recommended",    lambda r: "Yes" if r.get("is_cod_recommended") else "No", 14, FMT_TEXT),
    ("Rec. Priority",      "recommendation_priority",                         14, FMT_INT),
    ("AI Score Override",  "recommendation_score_override",                   16, FMT_PCT),
    # ── Legacy ───────────────────────────────────────────────────────────────
    ("Best Selling Scope", "best_selling_scope",                              16, FMT_TEXT),
    ("Sales Rank",         "sales_rank",                                      12, FMT_INT),
    # ── Bundle ───────────────────────────────────────────────────────────────
    ("Bundle Group",       "bundle_group",                                    16, FMT_TEXT),
    ("Bundle Discount %",  "bundle_discount_percent",                         16, FMT_PCT),
    # ── SAP ──────────────────────────────────────────────────────────────────
    ("Price Lock (SAP)",   lambda r: "Yes" if r.get("price_source_override") else "No", 14, FMT_TEXT),
    ("Last SAP Sync",      "last_synced_sap",                                 20, FMT_TEXT),
    # ── AI ───────────────────────────────────────────────────────────────────
    ("AI Score",           "ai_score",                                        12, FMT_DECIMAL),
    # ── Lifecycle ────────────────────────────────────────────────────────────
    ("Created At",         "created_at",                                      20, FMT_TEXT),
    ("Updated At",         "updated_at",                                      20, FMT_TEXT),
]

# Total column count — used to compute merged cell ranges
TOTAL_COLS: int = len(EXPORT_COLUMNS)
