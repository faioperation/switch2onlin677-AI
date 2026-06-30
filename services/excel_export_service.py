"""
services/excel_export_service.py
=================================
Enterprise-grade Excel export service for the Dhifaf Baghdad product catalog.

Performance architecture
------------------------
write_only=True workbook
    openpyxl serializes each row directly to an internal XML stream the moment
    ws.append() is called.  No in-memory cell grid is maintained.  Constant
    memory overhead regardless of dataset size.

SQLAlchemy yield_per(_CHUNK_SIZE)
    The DB cursor returns results in batches of 500 rows.  At most 500 product
    dicts exist in Python memory at any one time, even for a 100k-row export.

Single JOIN query
    Brand / Category / Subcategory names are resolved in the same query via
    LEFT OUTER JOINs.  Zero N+1 queries regardless of result count.

IQD rate loaded once
    The exchange rate JSON file is read once before the row loop begins.

Scaling guidance
----------------
< 50k rows   — current approach: BytesIO, returned synchronously, ~5-15 MB
50k–500k rows — swap BytesIO for NamedTemporaryFile + FileResponse
> 500k rows   — background job (Celery/APScheduler) + object storage presigned URL
"""
from __future__ import annotations

import io
import json
import logging
import os
from datetime import datetime, timezone
from typing import Any, Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from models import Brand, Category, Product, ProductSearchIndex, Subcategory
from utils.excel_styles import (
    ALIGN_CENTER,
    ALIGN_HEADER,
    ALIGN_LEFT,
    ALIGN_META,
    ALIGN_TITLE,
    BORDER_CELL,
    BORDER_HEADER,
    BORDER_NONE,
    EXPORT_COLUMNS,
    FILL_HEADER,
    FILL_IN_STOCK,
    FILL_LOW_STOCK,
    FILL_META,
    FILL_OUT_STOCK,
    FILL_ROW_EVEN,
    FILL_ROW_ODD,
    FILL_SEP,
    FILL_TITLE,
    FONT_DATA,
    FONT_HEADER,
    FONT_META_KEY,
    FONT_META_VAL,
    FONT_TITLE,
    FMT_TEXT,
    TOTAL_COLS,
)

logger = logging.getLogger(__name__)

_BAGHDAD_TZ = ZoneInfo("Asia/Baghdad")
_CHUNK_SIZE = 500                      # SQLAlchemy cursor batch size
_LAST_COL   = get_column_letter(TOTAL_COLS)   # "AI" for 35 columns
_DATA_START = 7                        # First data row (rows 1-6 are meta+header)

_RATE_FILE  = os.path.join(os.path.dirname(os.path.dirname(__file__)), "rate.json")

# ── Sort mapping (mirrors repositories/product.py) ────────────────────────────
_SORT_COLS: dict[str, tuple[str, str]] = {
    "name_asc":    ("item_name",     "ASC"),
    "name_desc":   ("item_name",     "DESC"),
    "price_asc":   ("price",         "ASC"),
    "price_desc":  ("price",         "DESC"),
    "stock_asc":   ("available_qty", "ASC"),
    "stock_desc":  ("available_qty", "DESC"),
    "created_asc": ("created_at",    "ASC"),
    "created_desc":("created_at",    "DESC"),
}
_SORT_ATTRS = {
    "item_name":     Product.item_name,
    "price":         Product.price,
    "available_qty": Product.available_qty,
    "created_at":    Product.created_at,
}


# ══════════════════════════════════════════════════════════════════════════════
# Internal helpers
# ══════════════════════════════════════════════════════════════════════════════

def _load_iqd_rate() -> float:
    try:
        with open(_RATE_FILE, "r", encoding="utf-8") as f:
            return float(json.load(f).get("iqd_rate", 1310))
    except Exception:
        return 1310.0


def _stock_status(qty: Any) -> str:
    return "in_stock" if qty and int(qty) > 0 else "out_of_stock"


def _mc(
    ws,
    value: Any = None,
    *,
    font=None,
    fill=None,
    alignment=None,
    border=None,
    num_fmt: Optional[str] = None,
) -> WriteOnlyCell:
    """Factory for a styled WriteOnlyCell."""
    cell = WriteOnlyCell(ws, value=value)
    if font:      cell.font          = font
    if fill:      cell.fill          = fill
    if alignment: cell.alignment     = alignment
    if border:    cell.border        = border
    if num_fmt:   cell.number_format = num_fmt
    return cell


def _row_to_dict(r: Any, rate: float) -> dict:
    """
    Serialize a raw SQLAlchemy multi-column named-tuple row into the standard
    product dict format used by EXPORT_COLUMNS.

    Deliberately self-contained (not delegating to ProductRepository._serialize)
    so the export service can be used independently of the repository layer.
    """
    price = float(r.price) if r.price is not None else 0.0
    tier  = getattr(r, "price_tier",    None)
    stat  = getattr(r, "product_status", None)
    rsco  = getattr(r, "recommendation_score_override", None)
    bdp   = getattr(r, "bundle_discount_percent",       None)
    ai_s  = getattr(r, "ai_score", None)

    return {
        "barcode":           r.barcode,
        "item_code":         r.item_code,
        "item_name":         r.item_name,
        "sap_product_id":    r.sap_product_id,
        "description":       r.description,
        "image_url":         r.image_url,
        "skin_type":         r.skin_type,
        "concerns":          list(r.concerns or []),
        "tags":              list(r.tags     or []),
        "price":             price,
        "price_iqd":         round(price * rate),
        "available_qty":     int(r.available_qty or 0),
        "stock_status":      _stock_status(r.available_qty),
        "price_tier":        getattr(tier,  "value", tier),
        "brand_family":      getattr(r, "brand_family", None),
        "product_status":    getattr(stat,  "value", stat),
        "is_best_selling":   bool(r.is_best_selling)      if r.is_best_selling      is not None else False,
        "is_new_arrival":    bool(getattr(r, "is_new_arrival",     None)) if getattr(r, "is_new_arrival",     None) is not None else False,
        "is_recommended":    bool(getattr(r, "is_recommended",     None)) if getattr(r, "is_recommended",     None) is not None else False,
        "is_cod_recommended":bool(getattr(r, "is_cod_recommended", None)) if getattr(r, "is_cod_recommended", None) is not None else False,
        "recommendation_priority":       getattr(r, "recommendation_priority", None),
        "recommendation_score_override": float(rsco) if rsco is not None else None,
        "best_selling_scope":getattr(r, "best_selling_scope", None),
        "sales_rank":        getattr(r, "sales_rank", None),
        "bundle_group":      getattr(r, "bundle_group", None),
        "bundle_discount_percent": float(bdp) if bdp is not None else None,
        "price_source_override": bool(getattr(r, "price_source_override", False)),
        "last_synced_sap":   r.last_synced_sap.isoformat() if r.last_synced_sap else None,
        "ai_score":          float(ai_s) if ai_s is not None else 0.0,
        "created_at":        r.created_at.isoformat() if r.created_at else None,
        "updated_at":        r.updated_at.isoformat() if r.updated_at else None,
        "brand":             {"id": r.brand_id,       "name": getattr(r, "brand_name",       None)},
        "category":          {"id": r.category_id,    "name": getattr(r, "category_name",    None)},
        "subcategory":       {"id": r.subcategory_id, "name": getattr(r, "subcategory_name", None)},
    }


def _build_filter_summary(
    q:               Optional[str],
    brand_id:        Optional[int],
    category_id:     Optional[int],
    subcategory_id:  Optional[int],
    is_best_selling: Optional[int],
    in_stock:        Optional[bool],
    min_price:       Optional[float],
    max_price:       Optional[float],
    product_status:  Optional[str],
) -> str:
    parts: list[str] = []
    if q:                          parts.append(f'Search="{q}"')
    if brand_id        is not None: parts.append(f"Brand ID={brand_id}")
    if category_id     is not None: parts.append(f"Category ID={category_id}")
    if subcategory_id  is not None: parts.append(f"Subcategory ID={subcategory_id}")
    if is_best_selling is not None:
        parts.append("Best Selling Only" if is_best_selling else "Non-Best-Selling")
    if in_stock:                    parts.append("In Stock Only")
    if min_price       is not None: parts.append(f"Min Price=${min_price:,.2f}")
    if max_price       is not None: parts.append(f"Max Price=${max_price:,.2f}")
    if product_status:              parts.append(f"Status={product_status}")
    return " | ".join(parts) if parts else "No filters — full catalog export"


# ══════════════════════════════════════════════════════════════════════════════
# Query builder
# ══════════════════════════════════════════════════════════════════════════════

def _build_export_query(
    db:              Session,
    *,
    q:               Optional[str]   = None,
    brand_id:        Optional[int]   = None,
    category_id:     Optional[int]   = None,
    subcategory_id:  Optional[int]   = None,
    is_best_selling: Optional[int]   = None,
    in_stock:        Optional[bool]  = None,
    min_price:       Optional[float] = None,
    max_price:       Optional[float] = None,
    product_status:  Optional[str]   = None,
    sort_by:         Optional[str]   = "created_desc",
):
    """
    Build the SQLAlchemy ORM query for the export.

    Mirrors the filter/sort logic from ProductRepository.list_products() but:
    - Selects every exportable column including ai_score (not in list_products)
    - Omits pagination (LIMIT / OFFSET)
    - Returns raw ORM row tuples suitable for yield_per() streaming

    Any filter or sort change in list_products should be mirrored here.
    """
    sort_col, sort_dir = _SORT_COLS.get(
        sort_by or "created_desc", ("created_at", "DESC")
    )

    select_cols = [
        # Core identity
        Product.barcode,            Product.item_code,          Product.item_name,
        Product.sap_product_id,
        # Display
        Product.description,        Product.image_url,
        # AI / Search
        Product.skin_type,          Product.concerns,           Product.tags,
        # Pricing
        Product.price,              Product.available_qty,
        # Classification
        Product.price_tier,         Product.brand_family,       Product.product_status,
        # Recommendation flags
        Product.is_best_selling,    Product.is_new_arrival,
        Product.is_recommended,     Product.is_cod_recommended,
        Product.recommendation_priority,
        Product.recommendation_score_override,
        # Legacy
        Product.best_selling_scope, Product.sales_rank,
        # Bundle
        Product.bundle_group,       Product.bundle_discount_percent,
        # SAP
        Product.price_source_override, Product.last_synced_sap,
        # AI
        Product.ai_score,
        # Lifecycle
        Product.created_at,         Product.updated_at,
        # FK ids (for dict assembly)
        Product.brand_id,           Product.category_id,        Product.subcategory_id,
        # Resolved names via JOIN
        Brand.name.label("brand_name"),
        Category.name.label("category_name"),
        Subcategory.name.label("subcategory_name"),
    ]

    query = (
        db.query(*select_cols)
        .join(Brand,       Brand.id       == Product.brand_id,       isouter=True)
        .join(Category,    Category.id    == Product.category_id,    isouter=True)
        .join(Subcategory, Subcategory.id == Product.subcategory_id, isouter=True)
        .filter(Product.deleted_at.is_(None))
    )

    # ── Full-text search (mirrors ProductRepository) ──────────────────────────
    if q:
        query = query.join(
            ProductSearchIndex,
            ProductSearchIndex.product_id == Product.barcode,
            isouter=True,
        )
        like = f"%{q.strip().lower()}%"
        query = query.filter(
            or_(
                ProductSearchIndex.search_text.ilike(like),
                func.to_tsvector("english", ProductSearchIndex.search_text).op("@@")(
                    func.plainto_tsquery("english", q.strip())
                ),
            )
        )

    # ── Column filters ────────────────────────────────────────────────────────
    if brand_id        is not None: query = query.filter(Product.brand_id        == brand_id)
    if category_id     is not None: query = query.filter(Product.category_id     == category_id)
    if subcategory_id  is not None: query = query.filter(Product.subcategory_id  == subcategory_id)
    if is_best_selling is not None: query = query.filter(Product.is_best_selling == is_best_selling)
    if in_stock:                    query = query.filter(Product.available_qty    >  0)
    if min_price       is not None: query = query.filter(Product.price            >= min_price)
    if max_price       is not None: query = query.filter(Product.price            <= max_price)
    if product_status  is not None: query = query.filter(Product.product_status   == product_status)

    # ── Sort ──────────────────────────────────────────────────────────────────
    attr = _SORT_ATTRS.get(sort_col, Product.created_at)
    query = query.order_by(attr.desc() if sort_dir == "DESC" else attr.asc())

    return query


# ══════════════════════════════════════════════════════════════════════════════
# Workbook section writers
# ══════════════════════════════════════════════════════════════════════════════

def _write_meta_section(
    ws,
    total: int,
    filter_summary: str,
    now: datetime,
) -> None:
    """Write rows 1-5: title banner, timestamp, count, filters, separator."""
    ts = now.astimezone(_BAGHDAD_TZ).strftime("%Y-%m-%d %I:%M %p") + " (Baghdad)"

    # Declare merged ranges BEFORE appending any rows.
    # In write_only mode, merge info is stored as XML metadata independently
    # of cell data and is safe to set at any point before save().
    for row_n in range(1, 6):
        try:
            ws.merge_cells(f"A{row_n}:{_LAST_COL}{row_n}")
        except (AttributeError, TypeError):
            pass  # Graceful degradation for environments where merge is unavailable

    def _merged(value: Any, font, fill, alignment=ALIGN_META) -> list:
        """Return a full-width row with `value` in the first cell."""
        first = _mc(ws, value, font=font, fill=fill, alignment=alignment)
        rest  = [_mc(ws, None, fill=fill) for _ in range(TOTAL_COLS - 1)]
        return [first] + rest

    # Row 1 — Company title (dark navy / gold)
    ws.append(_merged(
        "  Dhifaf Baghdad — Product Inventory Report",
        FONT_TITLE, FILL_TITLE, ALIGN_TITLE,
    ))

    # Row 2 — Generation timestamp
    ws.append(_merged(f"  Generated: {ts}", FONT_META_KEY, FILL_META))

    # Row 3 — Total record count
    ws.append(_merged(f"  Total Products: {total:,}", FONT_META_KEY, FILL_META))

    # Row 4 — Active filters
    ws.append(_merged(f"  Filters: {filter_summary}", FONT_META_VAL, FILL_META))

    # Row 5 — Visual separator
    ws.append([_mc(ws, None, fill=FILL_SEP) for _ in range(TOTAL_COLS)])


def _write_header_row(ws) -> None:
    """Write row 6: styled column header labels."""
    ws.append([
        _mc(ws, label,
            font=FONT_HEADER, fill=FILL_HEADER,
            alignment=ALIGN_HEADER, border=BORDER_HEADER)
        for (label, _, _, _) in EXPORT_COLUMNS
    ])


def _write_data_row(ws, seq: int, row_dict: dict, is_odd: bool) -> None:
    """
    Write one product data row with conditional stock-level coloring.

    Stock coloring hierarchy (highest priority first):
      out_of_stock    → FILL_OUT_STOCK  (soft red)
      available_qty 1–5 → FILL_LOW_STOCK  (amber warning)
      odd row         → FILL_ROW_ODD    (light blue stripe)
      even row        → FILL_ROW_EVEN   (white)
    """
    qty    = row_dict.get("available_qty", 0) or 0
    status = row_dict.get("stock_status", "out_of_stock")

    if status == "out_of_stock":
        row_fill = FILL_OUT_STOCK
    elif 0 < qty <= 5:
        row_fill = FILL_LOW_STOCK
    elif is_odd:
        row_fill = FILL_ROW_ODD
    else:
        row_fill = FILL_ROW_EVEN

    cells: list[WriteOnlyCell] = []
    for (_, key_or_fn, _, fmt) in EXPORT_COLUMNS:
        if key_or_fn == "__row_num__":
            value: Any = seq
        elif callable(key_or_fn):
            value = key_or_fn(row_dict)
        else:
            value = row_dict.get(key_or_fn)

        # Coerce None to empty string for text columns so Excel doesn't show "None"
        if value is None and fmt == FMT_TEXT:
            value = ""

        cells.append(_mc(ws, value,
                         font=FONT_DATA, fill=row_fill,
                         alignment=ALIGN_LEFT, border=BORDER_CELL,
                         num_fmt=fmt))
    ws.append(cells)


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════

def generate_export_bytes(
    db:              Session,
    *,
    q:               Optional[str]   = None,
    brand_id:        Optional[int]   = None,
    category_id:     Optional[int]   = None,
    subcategory_id:  Optional[int]   = None,
    is_best_selling: Optional[int]   = None,
    in_stock:        Optional[bool]  = None,
    min_price:       Optional[float] = None,
    max_price:       Optional[float] = None,
    product_status:  Optional[str]   = None,
    sort_by:         Optional[str]   = "created_desc",
) -> tuple[io.BytesIO, int]:
    """
    Generate the complete Excel workbook and return (buffer, total_count).

    The buffer is a seeked-to-zero BytesIO ready for streaming or saving.

    Parameters
    ----------
    db              : Active SQLAlchemy session (injected via FastAPI Depends)
    q               : Full-text search string
    brand_id        : Exact brand filter
    category_id     : Exact category filter
    subcategory_id  : Exact subcategory filter
    is_best_selling : 1=best sellers only, 0=non-best-sellers, None=all
    in_stock        : True=only products with available_qty > 0
    min_price       : Minimum price (USD) inclusive
    max_price       : Maximum price (USD) inclusive
    product_status  : "active" | "inactive" | "draft" | None (all)
    sort_by         : Sort key (see _SORT_COLS for valid values)

    Returns
    -------
    (BytesIO, int) — ready-to-read buffer and total matching product count
    """
    now  = datetime.now(tz=timezone.utc)
    rate = _load_iqd_rate()

    export_q = _build_export_query(
        db,
        q=q, brand_id=brand_id, category_id=category_id,
        subcategory_id=subcategory_id, is_best_selling=is_best_selling,
        in_stock=in_stock, min_price=min_price, max_price=max_price,
        product_status=product_status, sort_by=sort_by,
    )

    # Resolve total BEFORE starting the write-only stream (needed for row 3 header)
    total: int = db.execute(
        export_q.statement.with_only_columns(func.count()).order_by(None)
    ).scalar() or 0

    filter_summary = _build_filter_summary(
        q, brand_id, category_id, subcategory_id,
        is_best_selling, in_stock, min_price, max_price, product_status,
    )

    logger.info("product_export_start total=%d filters=%r", total, filter_summary)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Products")

    # Sheet-level properties (all safe in write_only mode)
    ws.sheet_properties.tabColor = "1E3A5F"   # Dark navy sheet tab
    ws.freeze_panes = f"A{_DATA_START}"        # Freeze rows 1-6 on scroll

    # Row heights (character units; 1 pt ≈ 0.75 px)
    ws.row_dimensions[1].height = 30   # Title banner
    ws.row_dimensions[2].height = 18   # Timestamp
    ws.row_dimensions[3].height = 18   # Total count
    ws.row_dimensions[4].height = 18   # Filters
    ws.row_dimensions[5].height = 6    # Separator
    ws.row_dimensions[6].height = 34   # Column headers

    # Column widths
    for col_idx, (_, _, width, _) in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Write sections ────────────────────────────────────────────────────────
    _write_meta_section(ws, total, filter_summary, now)
    _write_header_row(ws)

    rows_written = 0
    for raw_row in export_q.yield_per(_CHUNK_SIZE):
        rows_written += 1
        row_dict = _row_to_dict(raw_row, rate)
        _write_data_row(ws, rows_written, row_dict, is_odd=(rows_written % 2 == 1))

    logger.info("product_export_complete rows_written=%d", rows_written)

    # ── Serialize ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, total
